"""Integration test fixtures: FakeDrive, FakeNotifier, XLSX factories.

Real components: SqliteTracker, OpenpyxlExcelHandler, RowTransformer.
Faked components: Drive, Notifier, PathResolver, Lifecycle (Google API boundary).
"""

from __future__ import annotations

import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import pandas as pd
import pytest

from src.application.config import (
    AppConfig,
    DrivePathsConfig,
    EmailConfig,
    ExcelConfig,
    GoogleConfig,
    LoggingConfig,
    TrackingConfig,
)
from src.application.use_cases.consolidate_invoices import ConsolidateInvoicesUseCase
from src.infrastructure.excel_handler import OpenpyxlExcelHandler
from src.infrastructure.sqlite_tracker import SqliteTracker


# ── Column Constants ─────────────────────────────────────────────────

SOURCE_COLUMNS = [
    "N° Factura",
    "N° Referencia",
    "Transportista",
    "Fecha Factura",
    "Descripción",
    "Monto Neto",
    "IVA",
    "Monto Total",
    "Moneda",
]

CONSOLIDATED_COLUMNS = [
    "invoice_number",
    "reference_number",
    "carrier_name",
    "invoice_date",
    "description",
    "net_amount",
    "tax_amount",
    "total_amount",
    "currency",
    "source_file",
]


# ── XLSX Factory Functions ───────────────────────────────────────────


def create_source_xlsx(
    path: Path,
    rows: list[dict],
    sheet_name: str = "DETALLE FACTURACIÓN CONTENEDORE",
    startrow: int = 10,
) -> Path:
    """Create a source XLSX file with Spanish column names.

    Defaults to startrow=10 (row 11) to match OfficialFormatExtractor expectations.
    """
    df = pd.DataFrame(rows, columns=SOURCE_COLUMNS)
    df.to_excel(path, sheet_name=sheet_name, index=False, engine="openpyxl", startrow=startrow)
    return path


def create_mixed_format_source_xlsx(
    path: Path,
    fixed_cells: dict[str, str | int | None],
    tabular_rows: list[dict],
    sheet_name: str = "DETALLE FACTURACIÓN CONTENEDORE",
) -> Path:
    """Create a source XLSX file with mixed format (fixed cells + tabular data).

    Fixed cells layout:
    - B6: Empresa Transporte
    - B7: Fecha Emisión
    - B8: N° Factura
    - H6: Nave
    - H7: Puerto Embarque
    - F4: Aprobado por

    Tabular data starts at row 11.
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws["B6"] = fixed_cells.get("empresa_transporte")
    ws["B7"] = fixed_cells.get("fecha_emision")
    ws["B8"] = fixed_cells.get("numero_factura")
    ws["H6"] = fixed_cells.get("nave")
    ws["H7"] = fixed_cells.get("puerto_embarque")
    ws["F4"] = fixed_cells.get("aprobado_por")

    tabular_columns = [
        "Fecha Servicio",
        "Unidad",
        "Conductor",
        "Contenedor",
        "Patente Camión",
        "Patente Carro",
        "Órdenes de Embarque",
        "Plantas",
        "Guías de Despacho",
        "Cantidad Pallets",
        "Flete($)",
        "Underslung($)",
        "Planta Adicional ($)",
        "Retiro Cruzado ($)",
        "Porteo($)",
        "Hora Llegada Planta",
        "Hora Salida Planta",
        "Horas Sobre Estadía Planta",
        "Sobre Estadía Planta ($)",
        "Hora Llegada Puerto",
        "Hora Salida Puerto",
        "Horas Sobre Estadía Puerto",
        "Sobre Estadía Puerto ($)",
        "Fecha Gate In",
        "Fecha Gate Out",
        "Total Servicio ($)",
    ]

    for col_idx, col_name in enumerate(tabular_columns, start=1):
        ws.cell(row=11, column=col_idx, value=col_name)

    for row_idx, row_data in enumerate(tabular_rows, start=12):
        for col_idx, col_name in enumerate(tabular_columns, start=1):
            value = row_data.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(path)
    return path


def create_consolidated_xlsx(
    path: Path,
    rows: list[dict] | None = None,
    sheet_name: str = "Consolidado",
    header_row: int = 0,
) -> Path:
    """Create a consolidated XLSX file with standard column names."""
    if rows:
        df = pd.DataFrame(rows, columns=CONSOLIDATED_COLUMNS)
    else:
        df = pd.DataFrame(columns=CONSOLIDATED_COLUMNS)

    start_row = header_row - 1 if header_row > 0 else 0
    df.to_excel(
        path,
        sheet_name=sheet_name,
        index=False,
        engine="openpyxl",
        startrow=start_row,
    )
    return path


# ── Fake Implementations ────────────────────────────────────────────


class FakeDrive:
    """In-memory Drive that uses local files as backing storage.

    Implements the DriveRepository protocol with a local file registry.
    ``download_file`` copies from registry to destination;
    ``update_file`` copies destination back to registry so subsequent
    reads see the updated content.
    """

    def __init__(self) -> None:
        self._registry: dict[str, Path] = {}
        self._source_files: list[dict] = []
        self._find_results: dict[str, str] = {}
        self.calls: dict[str, list] = {
            "download": [],
            "upload": [],
            "update": [],
            "backup": [],
            "restore": [],
            "move": [],
        }

    # ── Setup helpers ─────────────────────────────────────────────

    def register(self, file_id: str, path: Path) -> None:
        """Register a local file to be served as a 'Drive file'."""
        self._registry[file_id] = path

    def set_source_files(self, files: list[dict]) -> None:
        """Set the list returned by ``list_source_files``."""
        self._source_files = files

    def set_find_result(self, filename: str, file_id: str) -> None:
        """Set the result for ``find_file_in_folder`` lookups."""
        self._find_results[filename] = file_id

    # ── DriveRepository Protocol ──────────────────────────────────

    def list_source_files(self, folder_id: str) -> list[dict]:
        return self._source_files

    def download_file(self, file_id: str, local_path: Path) -> Path:
        src = self._registry[file_id]
        shutil.copy2(src, local_path)
        self.calls["download"].append({"file_id": file_id, "local_path": local_path})
        return local_path

    def upload_file(self, local_path: Path, folder_id: str, file_name: str) -> str:
        uid = f"uploaded_{file_name}"
        dst = local_path.parent / f"_uploaded_{file_name}"
        shutil.copy2(local_path, dst)
        self._registry[uid] = dst
        self.calls["upload"].append(
            {"local_path": local_path, "folder_id": folder_id, "file_name": file_name}
        )
        return uid

    def create_backup(self, file_id: str, backup_name: str) -> str:
        backup_id = f"backup_{file_id}"
        src = self._registry[file_id]
        dst = src.parent / backup_name
        shutil.copy2(src, dst)
        self._registry[backup_id] = dst
        self.calls["backup"].append({"file_id": file_id, "backup_name": backup_name})
        return backup_id

    def restore_backup(self, backup_file_id: str, original_file_id: str) -> None:
        shutil.copy2(self._registry[backup_file_id], self._registry[original_file_id])
        self.calls["restore"].append(
            {"backup_file_id": backup_file_id, "original_file_id": original_file_id}
        )

    def update_file(self, file_id: str, local_path: Path) -> None:
        shutil.copy2(local_path, self._registry[file_id])
        self.calls["update"].append({"file_id": file_id, "local_path": local_path})

    def move_file(self, file_id: str, from_folder_id: str, to_folder_id: str) -> None:
        self.calls["move"].append({"file_id": file_id, "from": from_folder_id, "to": to_folder_id})

    def find_file_in_folder(self, folder_id: str, file_name: str) -> str | None:
        return self._find_results.get(file_name)

    def list_xlsx_in_folder(self, folder_id: str) -> list[dict]:
        return self._source_files


@dataclass
class FakeNotifier:
    """Captures notification calls for verification."""

    calls: list[dict] = field(default_factory=list)

    def send(
        self,
        subject: str,
        template_name: str,
        template_vars: dict[str, Any],
        recipients: list[str],
        cc: list[str] | None = None,
        bcc: list[str] | None = None,
        attachments: list[Path] | None = None,
    ) -> None:
        self.calls.append(
            {
                "subject": subject,
                "template_name": template_name,
                "template_vars": template_vars,
                "recipients": recipients,
                "cc": cc,
                "bcc": bcc,
                "attachments": attachments,
            }
        )


# ── Fixtures ─────────────────────────────────────────────────────────


@pytest.fixture
def app_config(tmp_path: Path) -> AppConfig:
    """Full AppConfig pointing to tmp_path for DB and credentials."""
    return AppConfig(
        google=GoogleConfig(credentials_path=str(tmp_path / "creds.json")),
        drive=DrivePathsConfig(
            source_path="Bot RPA/Tocornal/ETL Facturas",
            consolidated_path="Consolidado",
            consolidated_filename="consolidado.xlsx",
        ),
        excel=ExcelConfig(
            skip_schema_validation=True,
            source_sheet="DETALLE FACTURACIÓN CONTENEDORE",
        ),
        email=EmailConfig(
            sender="etl@smartbots.cl",
            to=("admin@smartbots.cl",),
            cc=(),
            bcc=(),
            subject_prefix="[Smartbots ETL]",
            templates={
                "success": "ETL_Consolidacion_Exito.html",
                "partial": "ETL_Consolidacion_Parcial.html",
                "error": "ETL_Consolidacion_Error.html",
                "empty": "ETL_Consolidacion_Vacio.html",
            },
        ),
        tracking=TrackingConfig(db_path=str(tmp_path / "tracking.db")),
        logging=LoggingConfig(),
    )


@pytest.fixture
def tracker(app_config: AppConfig) -> SqliteTracker:
    """Real SQLite tracker in tmp_path (no credentials needed)."""
    return SqliteTracker(app_config.tracking.db_path)


@pytest.fixture
def excel_handler() -> OpenpyxlExcelHandler:
    """Real Excel handler using openpyxl."""
    return OpenpyxlExcelHandler()


@pytest.fixture
def fake_drive() -> FakeDrive:
    return FakeDrive()


@pytest.fixture
def fake_notifier() -> FakeNotifier:
    return FakeNotifier()


@pytest.fixture
def path_resolver() -> MagicMock:
    """Mock PathResolver that maps known paths to deterministic folder IDs."""
    mock = MagicMock()

    def resolve(path: str) -> str:
        mapping = {
            "Bot RPA/Tocornal/ETL Facturas": "source_folder_id",
            "Consolidado": "consolidated_folder_id",
        }
        return mapping.get(path, f"folder_{path}")

    mock.resolve_path.side_effect = resolve
    return mock


@pytest.fixture
def lifecycle() -> MagicMock:
    """Mock FileLifecycleManager (move_to_in_process / move_to_backup)."""
    return MagicMock()


@pytest.fixture
def build_use_case(
    fake_drive: FakeDrive,
    excel_handler: OpenpyxlExcelHandler,
    fake_notifier: FakeNotifier,
    tracker: SqliteTracker,
    app_config: AppConfig,
    path_resolver: MagicMock,
    lifecycle: MagicMock,
):
    """Factory that builds a ConsolidateInvoicesUseCase with real + fake deps."""

    def _build(
        drive: FakeDrive | None = None,
        notifier: FakeNotifier | None = None,
    ) -> ConsolidateInvoicesUseCase:
        return ConsolidateInvoicesUseCase(
            drive=drive or fake_drive,
            reader=excel_handler,
            writer=excel_handler,
            notifier=notifier or fake_notifier,
            tracker=tracker,
            config=app_config,
            path_resolver=path_resolver,
            lifecycle=lifecycle,
        )

    return _build
