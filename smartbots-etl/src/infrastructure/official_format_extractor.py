"""Extractor oficial usando Pandas + Pydantic + Calamine (Rust backend)."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import pandas as pd
import structlog
from pydantic import BaseModel, Field, ValidationError

from src.application.config import ExcelConfig
from src.domain.entities import InvoiceRecord
from src.domain.exceptions import SchemaValidationError

logger = structlog.get_logger()


class FixedCells(BaseModel):
    """Celdas fijas del archivo origen."""

    empresa_transporte: str | None = Field(None, alias="C6")
    fecha_emision: str | None = Field(None, alias="G3")
    numero_factura: str | None = Field(None, alias="C8")
    nave: str | None = Field(None, alias="G6")
    puerto_embarque: str | None = Field(None, alias="G7")
    responsable: str | None = Field(None, alias="F4")

    @property
    def aprobado_por(self) -> str | None:
        """Limpia el prefijo 'Aprobado por: ' del responsable."""
        if self.responsable and isinstance(self.responsable, str):
            return self.responsable.replace("Aprobado por: ", "").strip()
        return self.responsable


class TabularRow(BaseModel):
    """Fila de datos tabulares desde fila 11."""

    fecha_servicio: str | None = Field(None, alias="Fecha Servicio")
    unidad: str | None = Field(None, alias="Unidad")
    conductor: str | None = Field(None, alias="Conductor")
    contenedor: str | None = Field(None, alias="Contenedor")
    patente_camion: str | None = Field(None, alias="Patente Camión")
    patente_carro: str | None = Field(None, alias="Patente Carro")
    ordenes_embarque: str | None = Field(None, alias="Órdenes de Embarque")
    plantas: str | None = Field(None, alias="Plantas")
    guias_despacho: str | int | float | None = Field(None, alias="Guías de Despacho")
    cantidad_pallets: Any | None = Field(None, alias="Cantidad Pallets")
    flete: Decimal | None = Field(None, alias="Flete($)")
    underslung: Decimal | None = Field(None, alias="Underslung($)")
    planta_adicional: Decimal | None = Field(None, alias="Planta Adicional ($)")
    retiro_cruzado: Decimal | None = Field(None, alias="Retiro Cruzado ($)")
    porteo: Decimal | None = Field(None, alias="Porteo($)")
    hora_llegada_planta: Any | None = Field(None, alias="Hora Llegada Planta")
    hora_salida_planta: Any | None = Field(None, alias="Hora Salida Planta")
    horas_sobre_estadia_planta: Any | None = Field(None, alias="Horas Sobre Estadía Planta")
    sobre_estadia_planta: Decimal | None = Field(None, alias="Sobre Estadía Planta ($)")
    hora_llegada_puerto: Any | None = Field(None, alias="Hora Llegada Puerto")
    hora_salida_puerto: Any | None = Field(None, alias="Hora Salida Puerto")
    horas_sobre_estadia_puerto: Any | None = Field(None, alias="Horas Sobre Estadía Puerto")
    sobre_estadia_puerto: Decimal | None = Field(None, alias="Sobre Estadía Puerto ($)")
    fecha_gate_in: str | None = Field(None, alias="Fecha Gate In")
    fecha_gate_out: str | None = Field(None, alias="Fecha Gate Out")
    total_servicio: Decimal = Field(Decimal("0"), alias="Total Servicio ($)")
    observaciones: str | None = Field(None, alias="Observaciones")


class OfficialFormatExtractor:
    """Extractor para formato oficial: celdas fijas + datos tabulares.

    Usa Pandas con engine='calamine' (si disponible) para máxima velocidad.
    """

    FIXED_CELLS = {
        "C6": "empresa_transporte",
        "G3": "fecha_emision",
        "C8": "numero_factura",
        "H6": "nave",
        "H7": "puerto_embarque",
        "F4": "responsable",  # Tiene "Aprobado por: " prefix
    }

    def __init__(self, config: ExcelConfig) -> None:
        self._config = config
        self._source_sheet = config.source_sheet
        self.validation_errors: list[dict] = []

    def extract(self, file_path: Path) -> list[InvoiceRecord]:
        """Extrae registros del archivo con formato oficial o formato simple tabular."""
        logger.debug(
            "debug_extract_start",
            file=str(file_path),
            sheet=self._source_sheet,
        )
        logger.debug(f"\n{'=' * 60}")
        logger.debug(f"DEBUG OfficialFormatExtractor.extract():")
        logger.debug(f"  → Archivo: {file_path.name}")
        logger.debug(f"  → Hoja configurada: {self._source_sheet}")

        self.validation_errors = []
        try:
            fixed = self._read_fixed_cells(file_path)
            logger.debug(f"  → Celdas fijas leídas:")
            logger.debug(f"      - Empresa Transporte: {fixed.empresa_transporte}")
            logger.debug(f"      - N° Factura: {fixed.numero_factura}")
            logger.debug(f"      - Nave: {fixed.nave}")

            is_mixed_format = (
                fixed.numero_factura is not None and fixed.empresa_transporte is not None
            )
            logger.debug(
                f"  → Formato detectado: {'MIXTO' if is_mixed_format else 'TABULAR SIMPLE'}"
            )

            if is_mixed_format:
                return self._extract_mixed_format(file_path, fixed)
            else:
                return self._extract_simple_tabular(file_path)

        except Exception as e:
            logger.error("extraction_failed", file=file_path.name, error=str(e))
            raise

    def _extract_mixed_format(self, file_path: Path, fixed: FixedCells) -> list[InvoiceRecord]:
        """Extrae registros usando formato mixto (celdas fijas + tabular)."""
        df = self._read_with_engine(file_path)

        logger.debug(
            "debug_dataframe_read",
            file=file_path.name,
            rows=len(df),
            columns=len(df.columns),
        )
        logger.debug(f"  → DataFrame leído: {len(df)} filas, {len(df.columns)} columnas")
        logger.debug(f"  → Columnas: {list(df.columns)[:10]}...")  # Primeras 10 columnas

        # Validar que las celdas clave tienen valores no nulos
        if (
            not fixed.numero_factura
            or not fixed.numero_factura.strip()
            or not fixed.empresa_transporte
            or not fixed.empresa_transporte.strip()
        ):
            raise SchemaValidationError(
                missing_columns=["N° Factura", "Empresa Transporte"],
                extra_columns=[],
            )

        records = []
        ordenes_column = "Órdenes de Embarque"

        logger.info(
            "debug_columns_check",
            file=file_path.name,
            columns=list(df.columns),
            ordenes_column_exists=ordenes_column in df.columns,
        )

        for idx, row in df.iterrows():
            try:
                if row.isna().all():
                    continue

                # Solo procesar filas que tengan valor en "Órdenes de Embarque"
                ordenes_val = row.get(ordenes_column) if ordenes_column in row.index else None
                if pd.isna(ordenes_val) or (
                    isinstance(ordenes_val, str) and not ordenes_val.strip()
                ):
                    continue

                row_values_str = " ".join([str(v).upper() for v in row.values if not pd.isna(v)])
                if any(kw in row_values_str for kw in ["NETO", "IVA", "TOTAL"]):
                    logger.debug("skipping_summary_row", row_index=idx, content=row_values_str)
                    continue

                row_dict = row.to_dict()
                row_dict = {k: (None if pd.isna(v) else v) for k, v in row_dict.items()}
                tabular = TabularRow.model_validate(row_dict)

                total = self._calculate_total(tabular)

                record = InvoiceRecord(
                    invoice_number=str(fixed.numero_factura),
                    reference_number=tabular.ordenes_embarque or "N/A",
                    carrier_name=str(fixed.empresa_transporte),
                    ship_name=str(fixed.nave) if fixed.nave else "",
                    dispatch_guides=str(tabular.guias_despacho) if tabular.guias_despacho else "",
                    invoice_date=self._parse_date(fixed.fecha_emision),
                    description=self._build_description(tabular, fixed),
                    net_amount=total,
                    tax_amount=Decimal("0"),
                    total_amount=total,
                    currency="CLP",
                    fecha_recepcion_digital="",
                    aprobado_por="",
                    estado_operaciones="",
                    fecha_aprobacion_operaciones="",
                    source_file=file_path.name,
                )
                records.append(record)

            except ValidationError as e:
                error_msg = f"Validation Error: {e!s}"
                self.validation_errors.append(
                    {"file": file_path.name, "row_index": int(idx), "error": error_msg}
                )
                logger.warning(
                    "row_validation_failed",
                    file=file_path.name,
                    row=int(idx),
                    errors=e.errors(),
                )
            except Exception as e:
                self.validation_errors.append(
                    {"file": file_path.name, "row_index": int(idx), "error": str(e)}
                )
                logger.warning(
                    "row_extraction_failed",
                    file=file_path.name,
                    row=int(idx),
                    error=str(e),
                )

        logger.info(
            "official_format_extracted",
            file=file_path.name,
            records_found=len(records),
            numero_factura=fixed.numero_factura,
            empresa=fixed.empresa_transporte,
        )

        if not records:
            logger.warning(
                "no_records_extracted",
                file=file_path.name,
                reason="No se encontraron filas con 'Órdenes de Embarque' con valor",
            )
            return []

        logger.debug("debug_extraction_mixed_complete", records=len(records))
        logger.debug(f"  → Registros extraídos (formato mixto): {len(records)}")
        logger.debug(f"{'=' * 60}\n")
        return records

    def _extract_simple_tabular(self, file_path: Path) -> list[InvoiceRecord]:
        """Extrae registros usando formato tabular simple (para tests/compatibilidad)."""
        df = self._read_tabular_data(file_path)

        records = []
        invoice_column = "N° Factura"

        for idx, row in df.iterrows():
            try:
                # Detectar fin de datos: si "N° Facturas" está vacío, detener extracción
                if invoice_column in row.index:
                    invoice_val = row.get(invoice_column)
                    if pd.isna(invoice_val) or (
                        isinstance(invoice_val, str) and not invoice_val.strip()
                    ):
                        logger.debug("debug_stop_extraction_empty_invoice", row_index=int(idx))
                        break

                if row.isna().all():
                    continue

                invoice_number = str(row.get("N° Factura", ""))
                if not invoice_number:
                    continue

                total_val = row.get("Monto Total", 0)
                net_val = row.get("Monto Neto", 0)
                tax_val = row.get("IVA", 0)

                def to_decimal(val):
                    if val is None or (isinstance(val, float) and (val != val)):
                        return Decimal("0")
                    return Decimal(str(val))

                total = to_decimal(total_val)
                net = to_decimal(net_val)
                tax = to_decimal(tax_val)

                record = InvoiceRecord(
                    invoice_number=invoice_number,
                    reference_number=str(row.get("N° Referencia")) or "N/A",
                    carrier_name=str(row.get("Transportista", "")),
                    ship_name=str(row.get("Nave", "")),
                    dispatch_guides=str(row.get("Guías de Despacho", "")),
                    invoice_date=self._parse_date(row.get("Fecha Factura")),
                    description=str(row.get("Descripción", "")),
                    net_amount=net,
                    tax_amount=tax,
                    total_amount=total,
                    currency=str(row.get("Moneda", "CLP")),
                    fecha_recepcion_digital=str(row.get("Fecha Recepción Digital", "")),
                    aprobado_por=str(row.get("Aprobado por:", "")),
                    estado_operaciones=str(row.get("Estado Operaciones", "")),
                    fecha_aprobacion_operaciones=str(row.get("Fecha Aprobación Operaciones", "")),
                    source_file=file_path.name,
                )
                records.append(record)

            except Exception as e:
                self.validation_errors.append(
                    {"file": file_path.name, "row_index": int(idx), "error": str(e)}
                )
                logger.warning(
                    "row_extraction_failed",
                    file=file_path.name,
                    row=int(idx),
                    error=str(e),
                )

        logger.info(
            "simple_tabular_extracted",
            file=file_path.name,
            records_found=len(records),
        )

        logger.debug("debug_extraction_simple_complete", records=len(records))
        logger.debug(f"  → Registros extraídos (formato simple): {len(records)}")
        logger.debug(f"{'=' * 60}\n")
        return records

    def _read_tabular_data(self, file_path: Path) -> pd.DataFrame:
        """Lee datos tabulares saltando 10 filas para coincidir con formato oficial."""
        try:
            import fastexcel

            reader = fastexcel.read_excel(file_path)
            df = reader.load_sheet_by_name(self._source_sheet).to_pandas()
            # Saltar 10 filas de encabezados
            df = df.iloc[10:] if len(df) > 10 else df
            # Usar la fila 11 como header
            df.columns = df.iloc[0].astype(str).tolist() if len(df) > 0 else df.columns.tolist()
            df = df[1:].reset_index(drop=True)
            return df
        except Exception:
            df = pd.read_excel(
                file_path,
                sheet_name=self._source_sheet,
                engine="openpyxl",
                header=None,
                skiprows=10,
            )
            # Asignar nombres de columnas desde fila 11
            df.columns = df.iloc[0].astype(str).tolist() if len(df) > 0 else df.columns.tolist()
            df = df[1:].reset_index(drop=True)
            return df

    def _read_with_engine(self, file_path: Path) -> pd.DataFrame:
        """Lee el archivo usando el mejor engine disponible."""
        try:
            import fastexcel

            reader = fastexcel.read_excel(file_path)
            df = reader.load_sheet_by_name(self._source_sheet).to_pandas()

            # Debug: mostrar contenido de filas clave para identificar estructura
            logger.info(
                "fastexcel_structure_debug",
                file=file_path.name,
                raw_rows=len(df),
                row_9=list(df.iloc[9])[:8] if len(df) > 9 else [],
                row_10=list(df.iloc[10])[:8] if len(df) > 10 else [],
                row_11=list(df.iloc[11])[:8] if len(df) > 11 else [],
            )

            # Buscar la fila que tiene "Órdenes de Embarque" como nombre de columna
            # NOTA: Buscamos específicamente "Órdenes de Embarque" porque "Puerto Embarque"
            # también contiene "Embarque" pero no es el header de la tabla de datos
            header_row_idx = None
            for idx in range(min(15, len(df))):
                row_values = [str(v) for v in df.iloc[idx] if pd.notna(v)]
                # Buscar específicamente "Órdenes de Embarque" o múltiples columnas de datos
                if any("Órdenes de Embarque" in v for v in row_values):
                    header_row_idx = idx
                    logger.info("header_found", row_index=idx, sample=row_values[:8])
                    break
                # También buscar si hay varias columnas conocidas (Fecha Servicio, Unidad, etc.)
                known_headers = {
                    "Fecha Servicio",
                    "Unidad",
                    "Conductor",
                    "Contenedor",
                    "Órdenes de Embarque",
                }
                if len(known_headers.intersection(set(row_values))) >= 3:
                    header_row_idx = idx
                    logger.info(
                        "header_found_by_known_columns", row_index=idx, sample=row_values[:8]
                    )
                    break

            if header_row_idx is not None:
                df = df.iloc[header_row_idx:]
                df.columns = df.iloc[0].astype(str).tolist()
                df = df[1:].reset_index(drop=True)
            elif len(df) > 10:
                df = df.iloc[10:]
                df.columns = df.iloc[0].astype(str).tolist()
                df = df[1:].reset_index(drop=True)
            else:
                df = df.reset_index(drop=True)

            logger.info(
                "fastexcel_processed",
                file=file_path.name,
                rows=len(df),
                columns=list(df.columns)[:10],
                ordenes_exists="Órdenes de Embarque" in df.columns,
            )

            logger.debug("used_fastexcel_engine")
            return df
        except ImportError:
            pass
        except Exception as e:
            logger.warning("fastexcel_failed", error=str(e))

        # Fallback a pandas con calamine engine si está instalado
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=self._source_sheet,
                engine="calamine",
                header=None,  # No usar header automático
                skiprows=10,  # Saltar filas 1-10 (encabezados fijos)
            )
            # Asignar nombres de columnas desde fila 11 (ahora es la primera fila del DF)
            df.columns = df.iloc[0].astype(str).tolist() if len(df) > 0 else df.columns.tolist()
            df = df[1:].reset_index(drop=True)
            logger.debug("used_pandas_calamine_engine")
            return df
        except (ImportError, ValueError):
            pass

        # Último fallback: openpyxl (más lento pero siempre disponible)
        df = pd.read_excel(
            file_path,
            sheet_name=self._source_sheet,
            engine="openpyxl",
            header=None,
            skiprows=10,
        )
        # Asignar nombres de columnas desde fila 11
        df.columns = df.iloc[0].astype(str).tolist() if len(df) > 0 else df.columns.tolist()
        df = df[1:].reset_index(drop=True)
        logger.debug("used_openpyxl_engine")
        return df

    def _read_fixed_cells(self, file_path: Path) -> FixedCells:
        """Lee las celdas fijas usando openpyxl."""
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)
        ws = wb[self._source_sheet]

        def _to_str(val: Any) -> str | None:
            if val is None:
                return None
            return str(val)

        cells = {
            "C6": _to_str(ws["C6"].value),
            "G3": _to_str(ws["G3"].value),
            "C8": _to_str(ws["C8"].value),
            "G6": _to_str(ws["H6"].value),
            "G7": _to_str(ws["H7"].value),
            "F4": _to_str(ws["F4"].value),
        }

        return FixedCells.model_validate(cells)

    def _calculate_total(self, row: TabularRow) -> Decimal:
        """Calcula el total sumando todos los componentes monetarios."""
        components = [
            row.flete or Decimal("0"),
            row.underslung or Decimal("0"),
            row.planta_adicional or Decimal("0"),
            row.retiro_cruzado or Decimal("0"),
            row.porteo or Decimal("0"),
            row.sobre_estadia_planta or Decimal("0"),
            row.sobre_estadia_puerto or Decimal("0"),
        ]

        # Si hay un total explícito, usarlo; si no, sumar componentes
        if row.total_servicio and row.total_servicio > 0:
            return row.total_servicio

        return sum(components, Decimal("0"))

    def _build_description(self, row: TabularRow, fixed: FixedCells) -> str:
        return str(row.observaciones) if row.observaciones else ""

    def _parse_date(self, value: Any) -> date:
        """Parsea fecha manejando strings, datetimes y timestamps."""
        if not value:
            raise ValueError("Date value is empty or None")

        if isinstance(value, (datetime, date)):
            if isinstance(value, datetime):
                return value.date()
            return value

        value_str = str(value).strip()

        formats = [
            "%d-%m-%Y",
            "%d-%m-%y",
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value_str, fmt).date()
            except ValueError:
                continue

        # Si no se puede parsear, lanzar ValueError
        raise ValueError(f"Formato de fecha inválido: {value!s}")
