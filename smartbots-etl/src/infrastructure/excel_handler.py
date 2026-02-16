from pathlib import Path

import openpyxl
import pandas as pd
import structlog

from openpyxl.styles import Alignment
from copy import copy

logger = structlog.get_logger()

_FALLBACK_SHEET = "Sheet1"

COLUMN_FORMATS = {
    "N° Factura": {"number_format": "0", "alignment": Alignment(horizontal="center")},
    "Empresa Transporte": {"alignment": Alignment(horizontal="center")},
    "Nave": {"alignment": Alignment(horizontal="center")},
    "Órdenes de Embarque": {"alignment": Alignment(horizontal="center")},
    "Guías de Despacho": {"number_format": "0", "alignment": Alignment(horizontal="right")},
    "Total Servicio ($)": {"number_format": '_ "$"* #,##0_ ;_ "$"* \-#,##0_ ;_ "$"* "-"_ ;_ @_ '},
    "Fecha Emisión": {"number_format": "dd/mm/yyyy", "alignment": Alignment(horizontal="center")},
    "Fecha Recepción Digital": {
        "number_format": "dd/mm/yyyy",
        "alignment": Alignment(horizontal="center"),
    },
    "Aprobado por:": {"alignment": Alignment(horizontal="center")},
    "Estado Operaciones": {"alignment": Alignment(horizontal="center")},
    "Fecha Aprobación Operaciones": {
        "number_format": "dd/mm/yyyy",
        "alignment": Alignment(horizontal="center"),
    },
    "Observaciones": {"alignment": Alignment(horizontal="left")},
}


class OpenpyxlExcelHandler:
    def read(
        self, file_path: Path, sheet_name: str = "Sheet1", header_row: int | None = None
    ) -> pd.DataFrame:
        actual_sheet = self._resolve_sheet(file_path, sheet_name)

        # pandas header is 0-indexed. header_row is 1-indexed (Excel).
        # header_row=None -> header=0 (default)
        header_arg = 0 if header_row is None else header_row - 1

        df = pd.read_excel(file_path, sheet_name=actual_sheet, header=header_arg, engine="openpyxl")
        logger.info(
            "excel_read",
            path=str(file_path),
            sheet=actual_sheet,
            requested_sheet=sheet_name,
            header_row=header_row,
            rows=len(df),
            columns=list(df.columns),
        )
        return df

    def write(
        self,
        df: pd.DataFrame,
        file_path: Path,
        sheet_name: str = "Sheet1",
        header_row: int = 0,
        data_start_row: int = 1,
    ) -> None:
        if not file_path.exists():
            df.to_excel(file_path, sheet_name=sheet_name, index=False, engine="openpyxl")
            logger.info("excel_created", path=str(file_path), rows=len(df))
            return

        wb = openpyxl.load_workbook(file_path)
        try:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                h_row = header_row if header_row > 0 else 1
                for col_idx, col_name in enumerate(df.columns, start=1):
                    ws.cell(row=h_row, column=col_idx, value=col_name)
                next_row = data_start_row if data_start_row > h_row else h_row + 1
                template_row = None
            else:
                ws = wb[sheet_name]
                next_row = self._find_next_empty_row(ws, min_row=data_start_row)
                template_row = next_row - 1 if next_row > data_start_row else None

            column_names = list(df.columns)

            for row_idx, row in enumerate(df.itertuples(index=False), start=next_row):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)

                    col_name = column_names[col_idx - 1] if col_idx <= len(column_names) else None

                    if col_name == "N° Factura" and value is not None:
                        try:
                            cell.value = int(float(str(value)))
                        except (ValueError, TypeError):
                            cell.value = value
                    else:
                        cell.value = value

                    if template_row and template_row >= data_start_row:
                        template_cell = ws.cell(row=template_row, column=col_idx)
                        if template_cell.has_style:
                            cell.font = copy(template_cell.font)
                            cell.border = copy(template_cell.border)
                            cell.fill = copy(template_cell.fill)
                            cell.protection = copy(template_cell.protection)

                    if col_name and col_name in COLUMN_FORMATS:
                        fmt = COLUMN_FORMATS[col_name]
                        if "number_format" in fmt:
                            cell.number_format = fmt["number_format"]
                        if "alignment" in fmt:
                            cell.alignment = fmt["alignment"]

            wb.save(file_path)
            logger.info(
                "excel_appended",
                path=str(file_path),
                rows_appended=len(df),
                start_row=next_row,
            )

        finally:
            wb.close()

    @staticmethod
    def _find_next_empty_row(ws, min_row: int) -> int:
        """Encuentra la siguiente fila vacía, respetando min_row."""
        if ws.max_row < min_row:
            return min_row

        # Buscar hacia arriba desde max_row hasta min_row
        for r in range(ws.max_row, min_row - 1, -1):
            if any(cell.value is not None for cell in ws[r]):
                return r + 1

        return min_row

    @staticmethod
    def _resolve_sheet(file_path: Path, requested: str) -> str:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheets = wb.sheetnames
            if requested in sheets:
                return requested
            if _FALLBACK_SHEET in sheets:
                logger.warning(
                    "sheet_fallback",
                    path=str(file_path),
                    requested=requested,
                    fallback=_FALLBACK_SHEET,
                )
                return _FALLBACK_SHEET
            raise ValueError(
                f"Sheet '{requested}' no encontrado en {file_path}. Sheets disponibles: {sheets}"
            )
        finally:
            wb.close()

    def validate_schema(
        self, df: pd.DataFrame, expected_columns: list[str]
    ) -> tuple[bool, list[str], list[str]]:
        actual = set(df.columns)
        expected = set(expected_columns)
        missing = sorted(expected - actual)
        extra = sorted(actual - expected)
        is_valid = len(missing) == 0

        logger.info(
            "schema_validation",
            is_valid=is_valid,
            missing=missing,
            extra=extra,
        )
        return is_valid, missing, extra
